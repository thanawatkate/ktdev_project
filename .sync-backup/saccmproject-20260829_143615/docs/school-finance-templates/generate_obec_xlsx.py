from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


OUTPUT_PATH = Path(__file__).resolve().parent / "OBEC-finance-templates.xlsx"


def style_header(ws, row=1):
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[row]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, widths):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def lock_all_cells(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.protection = Protection(locked=True)


def unlock_ranges(ws, ranges):
    for cell_range in ranges:
        for row in ws[cell_range]:
            for cell in row:
                cell.protection = Protection(locked=False)


def protect_sheet(ws, password="1234"):
    ws.protection.sheet = True
    ws.protection.password = password
    ws.protection.enable()
    ws.protection.sort = True
    ws.protection.autoFilter = True


def setup_print(ws, orientation="portrait", fit_width=1, fit_height=0):
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = fit_height
    ws.page_margins = PageMargins(left=0.35, right=0.35, top=0.5, bottom=0.5, header=0.3, footer=0.3)
    ws.print_options.horizontalCentered = True


def create_cover_sheet(wb):
    ws = wb.active
    ws.title = "หน้าปก"
    ws.merge_cells("A1:D1")
    ws["A1"] = "แบบประเมินมาตรฐานงานการเงินโรงเรียน (สพฐ.)"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    fields = [
        ("เลขที่เอกสาร", "OBEC-FIN-CHK-001"),
        ("เวอร์ชันแบบฟอร์ม", "v1.0"),
        ("หน่วยงานต้นสังกัด", "สำนักงานคณะกรรมการการศึกษาขั้นพื้นฐาน (สพฐ.)"),
        ("ชื่อโรงเรียน", ""),
        ("รหัสสถานศึกษา (ถ้ามี)", ""),
        ("ปีงบประมาณ", ""),
        ("ไตรมาส", ""),
        ("รอบประเมิน", ""),
        ("ผู้ประเมิน", ""),
        ("ตำแหน่ง", ""),
        ("วันที่ประเมิน", ""),
        ("หมายเหตุ", ""),
    ]
    start_row = 3
    for idx, (label, value) in enumerate(fields, start=start_row):
        ws[f"A{idx}"] = label
        ws[f"A{idx}"].font = Font(bold=True)
        ws[f"B{idx}"] = value
        ws.merge_cells(f"B{idx}:D{idx}")
        ws[f"B{idx}"].alignment = Alignment(wrap_text=True, vertical="top")

    budget_year_validation = DataValidation(type="list", formula1='"2568,2569,2570,2571,2572,2573,2574,2575"')
    ws.add_data_validation(budget_year_validation)
    budget_year_validation.add("B8")

    quarter_validation = DataValidation(type="list", formula1='"ไตรมาส 1,ไตรมาส 2,ไตรมาส 3,ไตรมาส 4,ทั้งปี"')
    ws.add_data_validation(quarter_validation)
    quarter_validation.add("B9")

    period_validation = DataValidation(type="list", formula1='"เดือน,ไตรมาส,ครึ่งปี,รายปี"')
    ws.add_data_validation(period_validation)
    period_validation.add("B10")

    date_validation = DataValidation(type="date", operator="between", formula1="DATE(2020,1,1)", formula2="DATE(2100,12,31)")
    date_validation.error = "กรุณากรอกวันที่ให้ถูกต้อง เช่น 30/04/2026"
    date_validation.errorTitle = "รูปแบบวันที่ไม่ถูกต้อง"
    ws.add_data_validation(date_validation)
    date_validation.add("B14")

    ws["A16"] = "คำแนะนำการใช้งาน"
    ws["A16"].font = Font(bold=True)
    ws.merge_cells("A17:D20")
    ws["A17"] = (
        "1) กรอกข้อมูลพื้นฐานในหน้าปกนี้\n"
        "2) ไปที่ชีต 'เช็กลิสต์มาตรฐาน' เพื่อให้คะแนน 0-2\n"
        "3) กรอกแผนแก้ไขในชีต 'แผนแก้ไข (CAP)'\n"
        "4) ตรวจผลรวมอัตโนมัติในชีต 'รายงานสรุป'"
    )
    ws["A17"].alignment = Alignment(wrap_text=True, vertical="top")

    auto_width(ws, {1: 30, 2: 36, 3: 10, 4: 10})
    setup_print(ws, orientation="portrait", fit_width=1, fit_height=1)

    lock_all_cells(ws)
    unlock_ranges(ws, ["B4:B14"])
    protect_sheet(ws)


def add_status_conditional(ws, cell_range):
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.conditional_formatting.add(
        cell_range, FormulaRule(formula=[f'{cell_range.split(":")[0]}="ผ่าน"'], fill=green)
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'{cell_range.split(":")[0]}="ต้องปรับปรุง"'], fill=yellow),
    )
    ws.conditional_formatting.add(
        cell_range, FormulaRule(formula=[f'{cell_range.split(":")[0]}="ไม่ผ่าน"'], fill=red)
    )


def add_percent_conditional(ws, cell_ref):
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(cell_ref, FormulaRule(formula=[f"{cell_ref}>=0.8"], fill=green))
    ws.conditional_formatting.add(
        cell_ref, FormulaRule(formula=[f"AND({cell_ref}>=0.6,{cell_ref}<0.8)"], fill=yellow)
    )
    ws.conditional_formatting.add(cell_ref, FormulaRule(formula=[f"{cell_ref}<0.6"], fill=red))


def build_checklist_sheet(wb):
    ws = wb.create_sheet("เช็กลิสต์มาตรฐาน")
    headers = [
        "หมวด",
        "รายการประเมิน",
        "คะแนน (0-2)",
        "น้ำหนัก",
        "คะแนนถ่วงน้ำหนัก",
        "สถานะ",
        "หมายเหตุ",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"

    items = [
        ("ธรรมาภิบาล", "มีคำสั่งมอบหมายงานการเงินเป็นลายลักษณ์อักษร"),
        ("ธรรมาภิบาล", "แยกหน้าที่ อนุมัติ/รับ/จ่าย/บันทึก ชัดเจน"),
        ("งบประมาณ", "มีแผนปฏิบัติการประจำปีเชื่อมโยงงบประมาณ"),
        ("งบประมาณ", "มีทะเบียนคุมงบรายโครงการเป็นปัจจุบัน"),
        ("งบประมาณ", "ตรวจสอบงบคงเหลือก่อนอนุมัติทุกครั้ง"),
        ("รับเงิน", "ออกเลขที่ใบเสร็จต่อเนื่อง ตรวจสอบย้อนหลังได้"),
        ("รับเงิน", "นำฝากเงินเข้าบัญชีภายในเวลาที่กำหนด"),
        ("รับเงิน", "กระทบยอดรายรับกับบัญชีธนาคารสม่ำเสมอ"),
        ("เบิกจ่าย", "เอกสารเบิกจ่ายครบก่อนอนุมัติจ่าย"),
        ("เบิกจ่าย", "จ่ายตรงตามวัตถุประสงค์และไม่เกินวงเงิน"),
        ("พัสดุ", "วิธีจัดซื้อจัดจ้างถูกต้องตามวงเงิน"),
        ("พัสดุ", "มีใบตรวจรับก่อนเบิกจ่ายทุกครั้ง"),
        ("บัญชี", "บันทึกรายการรับ-จ่ายเป็นปัจจุบัน"),
        ("บัญชี", "ปิดบัญชีและสรุปรายเดือนตรงเวลา"),
        ("ควบคุมภายใน", "กระทบยอดธนาคารทุกเดือนและมีผู้ตรวจทาน"),
        ("ควบคุมภายใน", "ตรวจนับเงินสดคงเหลือและจัดทำรายงาน"),
        ("รายงาน", "ส่งรายงานการเงินให้ผู้บริหารตามรอบเวลา"),
        ("รายงาน", "รายงานงบคงเหลือรายโครงการพร้อมใช้งาน"),
        ("ดิจิทัล", "กำหนดสิทธิ์ผู้ใช้งานตามบทบาท"),
        ("ดิจิทัล", "มี audit trail และแผนสำรองข้อมูล"),
    ]

    start_row = 2
    for i, (section, item) in enumerate(items, start=start_row):
        ws.append([section, item, "", 1, f"=IF(C{i}=\"\",\"\",C{i}*D{i})", "", ""])
        ws[f"C{i}"].alignment = Alignment(horizontal="center")
        ws[f"D{i}"].alignment = Alignment(horizontal="center")
        ws[f"E{i}"].number_format = "0.00"
        ws[f"F{i}"] = f'=IF(C{i}="","",IF(C{i}=2,"ผ่าน",IF(C{i}=1,"ต้องปรับปรุง","ไม่ผ่าน")))'

    score_validation = DataValidation(type="whole", operator="between", formula1="0", formula2="2")
    score_validation.error = "กรอกได้เฉพาะ 0, 1, หรือ 2"
    score_validation.errorTitle = "คะแนนไม่ถูกต้อง"
    score_validation.prompt = "คะแนน: 2=ผ่าน, 1=ต้องปรับปรุง, 0=ไม่ผ่าน"
    score_validation.promptTitle = "เกณฑ์การให้คะแนน"
    ws.add_data_validation(score_validation)
    score_validation.add(f"C{start_row}:C{start_row + len(items) - 1}")

    summary_start = start_row + len(items) + 1
    ws[f"A{summary_start}"] = "สรุปคะแนน"
    ws[f"A{summary_start}"].font = Font(bold=True)
    ws[f"B{summary_start}"] = "คะแนนรวม"
    ws[f"E{summary_start}"] = f"=SUM(E{start_row}:E{summary_start-2})"
    ws[f"B{summary_start+1}"] = "คะแนนเต็ม"
    ws[f"E{summary_start+1}"] = f"=SUM(D{start_row}:D{summary_start-2})*2"
    ws[f"B{summary_start+2}"] = "ร้อยละ"
    ws[f"E{summary_start+2}"] = f"=IF(E{summary_start+1}=0,\"\",E{summary_start}/E{summary_start+1})"
    ws[f"E{summary_start+2}"].number_format = "0.00%"
    ws[f"B{summary_start+3}"] = "ระดับผลประเมิน"
    ws[f"E{summary_start+3}"] = (
        f'=IF(E{summary_start+2}="","",IF(E{summary_start+2}>=0.8,"ดีมาก",'
        f'IF(E{summary_start+2}>=0.6,"พอใช้","เสี่ยงสูง")))'
    )

    style_header(ws)
    auto_width(ws, {1: 16, 2: 52, 3: 12, 4: 10, 5: 16, 6: 16, 7: 26})
    for r in range(2, summary_start + 4):
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    add_status_conditional(ws, f"F{start_row}:F{summary_start-2}")
    ws["C1"].comment = Comment("กรอกเฉพาะ 0, 1, 2", "ระบบ")
    setup_print(ws, orientation="landscape", fit_width=1, fit_height=0)

    lock_all_cells(ws)
    unlock_ranges(ws, [f"C{start_row}:C{summary_start-2}", f"G{start_row}:G{summary_start-2}"])
    protect_sheet(ws)


def build_cap_sheet(wb):
    ws = wb.create_sheet("แผนแก้ไข (CAP)")
    headers = [
        "ลำดับ",
        "ประเด็นที่ไม่ผ่าน/ต้องปรับปรุง",
        "สาเหตุราก (Root Cause)",
        "แผนแก้ไข",
        "ผู้รับผิดชอบ",
        "วันเริ่ม",
        "กำหนดเสร็จ",
        "หลักฐานที่ต้องมี",
        "% ความคืบหน้า",
        "สถานะ",
        "หมายเหตุ",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"
    for i in range(2, 12):
        ws.append([i - 1, "", "", "", "", "", "", "", "", "ยังไม่เริ่ม", ""])
        ws[f"I{i}"].number_format = "0%"

    status_validation = DataValidation(
        type="list",
        formula1='"ยังไม่เริ่ม,กำลังดำเนินการ,รอตรวจทาน,เสร็จ"',
    )
    ws.add_data_validation(status_validation)
    status_validation.add("J2:J11")

    progress_validation = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")
    progress_validation.error = "กรอกค่าระหว่าง 0 ถึง 1 (เช่น 0.25 = 25%)"
    progress_validation.errorTitle = "ค่าความคืบหน้าไม่ถูกต้อง"
    ws.add_data_validation(progress_validation)
    progress_validation.add("I2:I11")

    ws["A13"] = "สรุป"
    ws["B13"] = "จำนวนรายการ"
    ws["C13"] = "=COUNTA(A2:A11)"
    ws["B14"] = "เสร็จแล้ว"
    ws["C14"] = '=COUNTIF(J2:J11,"เสร็จ")'
    ws["B15"] = "อัตราเสร็จสิ้น"
    ws["C15"] = '=IF(C13=0,"",C14/C13)'
    ws["C15"].number_format = "0.00%"

    style_header(ws)
    auto_width(ws, {1: 8, 2: 34, 3: 26, 4: 30, 5: 16, 6: 12, 7: 12, 8: 24, 9: 12, 10: 14, 11: 20})
    for r in range(2, 16):
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"D{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    setup_print(ws, orientation="landscape", fit_width=1, fit_height=0)

    lock_all_cells(ws)
    unlock_ranges(ws, ["B2:K11"])
    protect_sheet(ws)


def build_report_sheet(wb):
    ws = wb.create_sheet("รายงานสรุป")
    rows = [
        ("หัวข้อ", "ค่า"),
        ("ชื่อโรงเรียน", "='หน้าปก'!B6"),
        ("สังกัด", "สพฐ."),
        ("รอบประเมิน", "='หน้าปก'!B10"),
        ("ผู้ประเมิน", "='หน้าปก'!B11"),
        ("ปีงบประมาณ", "='หน้าปก'!B8"),
        ("ไตรมาส", "='หน้าปก'!B9"),
        ("คะแนนรวม", "='เช็กลิสต์มาตรฐาน'!E23"),
        ("คะแนนเต็ม", "='เช็กลิสต์มาตรฐาน'!E24"),
        ("ร้อยละ", "='เช็กลิสต์มาตรฐาน'!E25"),
        ("ระดับผลประเมิน", "='เช็กลิสต์มาตรฐาน'!E26"),
        ("ความคืบหน้า CAP", "='แผนแก้ไข (CAP)'!C15"),
        ("จุดแข็ง 1", ""),
        ("จุดแข็ง 2", ""),
        ("จุดแข็ง 3", ""),
        ("จุดที่ต้องแก้ 1", ""),
        ("จุดที่ต้องแก้ 2", ""),
        ("จุดที่ต้องแก้ 3", ""),
        ("แผนเร่งด่วน สัปดาห์ 1-2", ""),
        ("แผนเร่งด่วน สัปดาห์ 3-4", ""),
        ("แผนเร่งด่วน เดือนที่ 2", ""),
        ("ความเสี่ยงหากไม่แก้ไข", ""),
        ("คำสั่งผู้อำนวยการ", ""),
        ("วันที่อนุมัติ", ""),
        ("ผู้จัดทำ", ""),
        ("ตำแหน่ง", ""),
    ]
    for row in rows:
        ws.append(list(row))

    signature_start = len(rows) + 3
    ws[f"A{signature_start}"] = "ลงนาม"
    ws[f"A{signature_start}"].font = Font(bold=True)

    ws[f"A{signature_start+1}"] = "ผู้จัดทำ"
    ws[f"B{signature_start+1}"] = "...................................................."
    ws[f"A{signature_start+2}"] = "วันที่"
    ws[f"B{signature_start+2}"] = ""

    ws[f"A{signature_start+4}"] = "ผู้ตรวจทาน"
    ws[f"B{signature_start+4}"] = "...................................................."
    ws[f"A{signature_start+5}"] = "วันที่"
    ws[f"B{signature_start+5}"] = ""

    ws[f"A{signature_start+7}"] = "ผู้อำนวยการโรงเรียน"
    ws[f"B{signature_start+7}"] = "...................................................."
    ws[f"A{signature_start+8}"] = "วันที่"
    ws[f"B{signature_start+8}"] = ""

    ws[f"A{signature_start+10}"] = "ความเห็นผู้ตรวจสอบ/ข้อเสนอแนะ"
    ws[f"A{signature_start+10}"].font = Font(bold=True)
    ws.merge_cells(f"B{signature_start+10}:D{signature_start+14}")
    ws[f"B{signature_start+10}"] = ""
    ws[f"B{signature_start+10}"].alignment = Alignment(wrap_text=True, vertical="top")

    style_header(ws)
    auto_width(ws, {1: 30, 2: 68})
    for r in range(2, signature_start + 15):
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B10"].number_format = "0.00%"
    ws["B12"].number_format = "0.00%"
    ws[f"B{signature_start+2}"].number_format = "dd/mm/yyyy"
    ws[f"B{signature_start+5}"].number_format = "dd/mm/yyyy"
    ws[f"B{signature_start+8}"].number_format = "dd/mm/yyyy"
    add_percent_conditional(ws, "B10")
    add_percent_conditional(ws, "B12")

    date_validation = DataValidation(type="date", operator="between", formula1="DATE(2020,1,1)", formula2="DATE(2100,12,31)")
    ws.add_data_validation(date_validation)
    date_validation.add(f"B{signature_start+2}")
    date_validation.add(f"B{signature_start+5}")
    date_validation.add(f"B{signature_start+8}")
    setup_print(ws, orientation="portrait", fit_width=1, fit_height=1)

    lock_all_cells(ws)
    unlock_ranges(
        ws,
        [
            "B13:B26",
            f"B{signature_start+2}:B{signature_start+2}",
            f"B{signature_start+5}:B{signature_start+5}",
            f"B{signature_start+8}:B{signature_start+8}",
            f"B{signature_start+10}:B{signature_start+14}",
        ],
    )
    protect_sheet(ws)


def build_dashboard_sheet(wb):
    ws = wb.create_sheet("Dashboard")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Dashboard สรุปความพร้อมการเงินโรงเรียน"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = "ชื่อโรงเรียน"
    ws["B3"] = "='หน้าปก'!B6"
    ws["A4"] = "ปีงบประมาณ"
    ws["B4"] = "='หน้าปก'!B8"
    ws["A5"] = "ไตรมาส"
    ws["B5"] = "='หน้าปก'!B9"

    ws["A7"] = "ตัวชี้วัดหลัก"
    ws["A7"].font = Font(bold=True)
    ws["A8"] = "ร้อยละผลประเมินเช็กลิสต์"
    ws["B8"] = "='เช็กลิสต์มาตรฐาน'!E25"
    ws["A9"] = "ระดับผลประเมิน"
    ws["B9"] = "='เช็กลิสต์มาตรฐาน'!E26"
    ws["A10"] = "ความคืบหน้า CAP"
    ws["B10"] = "='แผนแก้ไข (CAP)'!C15"
    ws["A11"] = "รายการ CAP ทั้งหมด"
    ws["B11"] = "='แผนแก้ไข (CAP)'!C13"
    ws["A12"] = "รายการ CAP ที่เสร็จ"
    ws["B12"] = "='แผนแก้ไข (CAP)'!C14"

    ws["A14"] = "สรุปสถานะรวม"
    ws["B14"] = '=IF(B8>=0.8,IF(B10>=0.8,"พร้อมรับการตรวจ","ต้องเร่งติดตาม CAP"),IF(B8>=0.6,"พอใช้ ควรเร่งปรับปรุง","เสี่ยงสูง ต้องแก้ไขด่วน"))'
    ws.merge_cells("B14:D14")
    ws["B14"].alignment = Alignment(wrap_text=True, vertical="center")

    ws["B8"].number_format = "0.00%"
    ws["B10"].number_format = "0.00%"
    add_percent_conditional(ws, "B8")
    add_percent_conditional(ws, "B10")

    auto_width(ws, {1: 30, 2: 35, 3: 18, 4: 18})
    setup_print(ws, orientation="portrait", fit_width=1, fit_height=1)

    lock_all_cells(ws)
    protect_sheet(ws)


def main():
    wb = Workbook()
    create_cover_sheet(wb)
    build_checklist_sheet(wb)
    build_cap_sheet(wb)
    build_report_sheet(wb)
    build_dashboard_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
